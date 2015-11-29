from openpyxl import Workbook
from openpyxl import load_workbook
import csv
import sys

def get_workbook(xlsx):
    ''' opens the workbook. needs an xlsx as an input
    returns an open workbook'''
    return load_workbook(xlsx)

def get_worksheet_list(wb):
    ''' creates a list of all of the worksheets found in the workbook, needs workbook as an input
    returns the list name and the worksheet object. skips any worksheets called overview'''
    ws_list = []
    worksheets = []

    for ws in wb.get_sheet_names():
        if ws.lower() == 'overview': continue
        ws_list.append(ws)
        worksheets.append(wb.get_sheet_by_name(ws))

    #worksheet_info = zip(ws_list, worksheets)
    return worksheets

def get_region(ws, loc = 'B1'):
    '''Gives you the region the worksheet is referring to.
    Assumptions:
    -region can be found in B1 as the first word  
    '''
    region = ws[loc].value.split(' ')[0]
    return region


def get_flow(ws):
    '''Gives you a list with flow values 
    Assumptions:
    -flow is either Income or Expenditure
    -flow is found in the C column
    '''    
    last_row = ws.max_row
    first_row = ws.min_row
    
    C_first = 'C'+ `first_row` 
    C_last = 'C'+ `last_row`
    
    flow = ['Income','Expenditure']
    the_C_range = ws[C_first:C_last] #the range of the elements in column C
    flow_cells = [] #list of tuples of (cell, flow)
    
    #create list of tuples of the (cell, value) pairs in all of the C column that have the flow as a value
    for row in the_C_range:
        for cell in row:
            if type(cell.value) == unicode and cell.value in flow:
                #flow_cells.append((cell, cell.row,cell.value))
                flow_cells.append(cell.value)
    #return the region, flow, entity and budget item info
    return flow_cells


def get_entity(ws):
    '''Gives you a list of entity values
    
    Assumptions:
    -entities are either (High Court, Advocate General, Auditor General) (In million kyats),
        (Ministries, Administrative Departments, Municipals) (in Million kyats), or
        (State Owned Enterprises) (in million kyats)
    -entities are found in the B column 
    '''
    last_row = ws.max_row
    first_row = ws.min_row
    
    B_first = 'B'+ `first_row`
    B_last = 'B'+ `last_row`

    
    entity = ['(High Court, Advocate General, Auditor General) (In million kyats)',
        '(Ministries, Administrative Departments, Municipals) (in Million kyats)',
        '(State Owned Enterprises) (in million kyats)']   
    
    the_B_range = ws[B_first:B_last] #the range of elements in column Bmn C
    the_cells = []
    
    #create a list of tuples of the (cell, value) pairs in all of the B column --everything
    for row in the_B_range:
        for cell in row:
            the_cells.append((cell,cell.value))
    
    #final lists we care about
    entity_cells = [] #list of tuples of (cell, entity)
    

    #create the list of cells with the entity values
    #create the list of tuples with cell, entity values pair
    for cell, value in the_cells:
        for ent in entity:
            if value != None and ent.lower() in value.lower(): #lower so that you capture everything
                #entity_cells.append((cell, cell.row, ent))
                entity_cells.append(ent)
    #return the region, flow, entity and budget item info
    return entity_cells


def get_sections(ws):
    '''Gives you a list of tuples with the row number of the start and ends of the sections by giving you 
    the cells of 'budget item' and 'total' 
    
    Assumptions:
    -budget items are between the rows labeled 'budget item' and 'total'
    -budget items are in column B
    '''
    last_row = ws.max_row
    first_row = ws.min_row
    
    B_first = 'B'+ `first_row`
    B_last = 'B'+ `last_row`
    
    the_B_range = ws[B_first:B_last] #the range of elements in column B
    
    budget_item_cells = [] #the cells that contain "budget item" aka the start of the section
    total_cells = [] #the cells that contain "total" aka the end of the section
    
    #creates list of 'budget item' cells
    #creates list of 'total' cells
    for row in the_B_range:
        for cell in row:
            if not cell.value: continue
            if cell.value.lower() == 'budget item':
                #print cell.value
                #budget_item_cells.append((cell, cell.row, cell.value)) 
                budget_item_cells.append(cell.row) 
            if cell.value.lower() == 'total':
                #print cell.value
                #total_cells.append((cell, cell.row, cell.value))
                total_cells.append(cell.row)
    
    #creates list of tuple with section ranges in each tuple/element of the list
    sections = zip(budget_item_cells, total_cells)
    
    #return the region, flow, entity and budget item info
    return sections


def get_source_cells(row):
    """Gives you the sources start and end locations

    Assumptions:
    -Sources are in the same line as "Budget Item" from each section
    """
    start_source = 'C' + `row`
    end_source = 'K' + `row`
    the_range = start_source, end_source
    
    return the_range

def get_budget_cells(row):
    """"Gives you the budget start and end locations

    Assumptions:
    -Budgets are in B column"""
    the_row = 'B' + `row`
    
    return the_row  

def get_value_cells(start_row, end_row):
    """Gives you the value start and end locations

    Assumptions:
    -Values of interest encompass the data between budget items and total in the B column
     (but not including those rows), up to the Total column in column L (but not including those rows)
    """
    start_source = 'C' + `start_row`
    end_source = 'K' + `end_row`
    the_range = start_source, end_source
    
    return the_range


def get_items(ws,c1,c2):
    '''Creates a list of the budget items from given worksheet, and area of budget items
    ws = the worksheet
    c1 the location of the first item (cell)
    c2 is the location of the last item (cell)

    Returns a list of items--- used to get budget items, sources, and values
    '''
    the_items = ws[c1:c2]

    items_list = []
    items_cell = []
    
    for row in the_items:
        for cell in row:
            items_list.append(cell.value)
            items_cell.append(cell)
    return items_list



def grab_info(ws, section_list):
    """Creates 3 lists: list of budgets, list of sources, list of values
        each index in the list corresponds to a table; budgets[0], sources[0], values[0] 
        are all the data for the first table in the worksheet

        Returns 3 lists:
        [0] is budgets
        [1] is sources
        [2] is values
    """
    budgets = []
    sources = []
    values = []
    
    #budgets do not include the 'Budget Item' row or the 'Total' row
    #values do not include the 'Budget Item' row or the 'Total' row
    for item in section_list:
        budgets.append(get_items(ws, get_budget_cells(item[0]+1), get_budget_cells(item[1]-1)))
        sources.append(get_items(ws, get_source_cells(item[0])[0], get_source_cells(item[0])[1]))
        values.append(get_items(ws, get_value_cells(item[0]+1, item[1])[0], get_value_cells(item[0]+1, item[1]-1)[1]))
    
    #Where the sources have no header value/no source value call it 'Delete Cell' for future deletion
    #needed because of how values are received 
    for lst in sources:
        for item in lst:
            #print item
            if item == None:
                idx = lst.index(item)
                item = u'Delete Cell'
                lst[idx] = item

    #where there is no value, call it NULL so there is no confusion and can be deleted out if requested
    for lst in values:
        for item in lst:
            if item == None:
                idx = lst.index(item)
                item = u'NULL' 
                lst[idx] = item
    
    return budgets, sources, values
    
    
def compose_data(region, flows, entities, budgets, sources, values):
    """Create the rows to be pushed to the csv file in the order of region, flow, entity, budget, source, value
    Returns each row as a list inside a list
    """
    data_part1 = []
    data_part2 = []

    for table_number, flow in enumerate(flows):
        for budget in budgets[table_number]:
            for source in sources[table_number]:
                data_part1.append([region, flow, entities[table_number], budget, source])   
        for value in values[table_number]:
            data_part2.append(value)
            
    zip_data = zip(data_part1,data_part2)
    
    full_data_list = []
    
    for lst, val in zip_data:
        lst.append(val)
        full_data_list.append(lst)
    
    return full_data_list


def clean_up(data):
    """Cleans up the data: deletes the rows that have 'Delete Cell' as a source
        as configured in grab_info()"""
    for idx, item in enumerate(data):
        if item[4] == 'Delete Cell':
            data.pop(idx)
    return data


def get_worksheet_data(worksheet):
    regions = get_region(worksheet)
    flows = get_flow(worksheet)
    entities = get_entity(worksheet)
    sections = get_sections(worksheet)
    budgets = grab_info(worksheet, sections)[0]
    sources = grab_info(worksheet, sections)[1]
    values = grab_info(worksheet, sections)[2]

    return regions, flows, entities, sections, budgets, sources, values


def map_xlsx_to_csv(clean_data):
    """Push rows of data to a file called myanmar_clean_data.csv  """
    with open('myanmar_clean_data.csv', 'wb') as csvfile:
        write = csv.writer(csvfile, delimiter=',')
        for row in clean_data:
            write.writerow(row)


def generate_csv_files(xlsx):
    """Putting all functions together and iterating through all worksheets"""
    csv_contents = []

    #open xlsx file
    workbook = get_workbook(xlsx)

    #get worksheets in xlsx
    worksheet_list = get_worksheet_list(workbook)
    
    for worksheet in worksheet_list:
        region, flow, entity, section, budget, source, values = get_worksheet_data(worksheet)
        clean_data = clean_up(compose_data(region, flow, entity, budget, source, values))
        csv_contents += clean_data

    #return csv_contents
    map_xlsx_to_csv(csv_contents)

if __name__ == '__main__':
    filename = None
    error = 'No file name provided. Please specify a .xlsx file.'
    if len(sys.argv) > 1:
        filename = sys.argv[1]
        if '.xlsx' not in filename:
            error = 'The argument "%s" is not a .xlsx file. Please provide one.' % str(filename)
            filename = None

    if filename:
        generate_csv_files(filename)
        print 'All done!'
    else:
        print error